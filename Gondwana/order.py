#!/usr/bin/env python
# encoding: utf-8

import os
from datetime import datetime
from flask import Blueprint, render_template, redirect, url_for, request, current_app, flash, send_file
from Gondwana.model import *
from Gondwana.helper import *
from pycscart import Cscart
from sqlalchemy import or_
from werkzeug.utils import secure_filename

bp = Blueprint('order', __name__, url_prefix='/order')


# /order/upload/tracking
@bp.route('/upload/tracking', methods=('POST', ))
def order_upload_tracking():
    def allowed_file(filename):
        return '.' in filename and filename.rsplit('.', 1)[1].lower() in current_app.config['ALLOWED_EXTENSIONS']

    if 'file' not in request.files:
        flash('No file part.', 'error')
        return redirect(url_for('order.order_index'))

    # get file
    file = request.files['file']
    if file.filename == '':
        flash('No selected file.', 'error')
        return redirect(url_for('order.order_index'))

    if file and allowed_file(file.filename): # save file
        filename = secure_filename(file.filename)
        file.save(os.path.join(current_app.config['UPLOAD_FOLDER'], filename))
        return redirect(url_for('order.order_parse_tracking', filename=filename))

# /order/parse/tracking/<string: filename>
@bp.route('/parse/tracking/<string:filename>', methods=('GET', ))
def order_parse_tracking(filename: str):
    filename = os.path.join(current_app.config['UPLOAD_FOLDER'], filename)

    from openpyxl import load_workbook

    wb = load_workbook(filename)
    ws = wb.active

    line = 2
    interval = 13
    while True:
        # get order id and tracking NO.
        order_id = ws['A'+str(line)].value

        # break if no new lines
        if order_id is None:
            break

        # get order
        order = Order.query.filter(Order.order_id == order_id).first()

        # get tracking NO.
        order.tracking_no = str(ws['L'+str(line)].value)

        # parse tracking NO. to add order.shipping_method
        order.shipping_method = generate_shipping_method(order.tracking_no)

        # parse tracking NO. to add to order.tracking_no or order.references_no
        if order.shipping_method == 'DHL': # DHL use references_no
            order.references_no = order.tracking_no
            order.tracking_no = None
        # add now to order.ship_time
        order.ship_time = datetime.now()

        # change order.status to shipped
        # send mail to customer to notify that package has been sent
        order.status = 'A' # shipped status // update order status automatically

        # update order
        db.session.commit()

        # update line number
        line += interval

    # remove uploaded file
    remove_file(filename)

    flash('Tracking NO. imported.', 'success')
    return redirect(url_for('order.order_index'))

# /order/download
@bp.route('/download', methods=('POST', ))
def order_download():
    download = request.form.get('download')

    if download:  # download order
        order_ids = request.form.get('order_ids')
        if order_ids:
            order_ids = [int(x) for x in order_ids.split(',')]  # change string to integer
        # https://stackoverflow.com/questions/15267755/query-for-multiple-values-at-once
        batching_orders = Order.query.filter(Order.id.in_(order_ids))

    import xlsxwriter

    xlsx_filename = os.path.join(current_app.instance_path,
                                 current_app.config['ORDER_DOWNLOAD_FILENAME'])
    workbook = xlsxwriter.Workbook(xlsx_filename)
    worksheet = workbook.add_worksheet()

    # generate
    first_download_time = datetime.now()
    titles = [
        'Order ID', 'Download Time', '', '', '', '', '', '', 'supply',
        'supply price', 'Express', 'Tracking No.', 'Express Price',
        'Ship Time', 'Product memo', 'Order memo'
    ]

    merge_format = workbook.add_format({
        'align': 'center',
        'valign': 'vcenter'
    })

    def merge_lines_data(worksheet, first_line, col, interval, data,
                         merge_format):
        end_line = first_line + interval - 1
        worksheet.merge_range(first_line, col, end_line, col, data,
                              merge_format)

    def format_first_download_time(first_download_time):
        return first_download_time.strftime('%Y.%m.%d')

    worksheet.merge_range('C1:H1', '')
    for i in range(len(titles)):
        worksheet.write(0, i, titles[i])

    start = 1
    interval = 13
    end = start + interval - 1
    for order in batching_orders:
        for item_id, product in order.products.items():
            # Order ID
            merge_lines_data(worksheet, start, 0, interval, order.order_id,
                         merge_format)
            # Download Time
            if order.first_download_time:
                time = format_first_download_time(order.first_download_time)
            else:
                time = format_first_download_time(first_download_time)
                order.first_download_time = first_download_time  # update first_download_time
            merge_lines_data(worksheet, start, 1, interval, time, merge_format)

            # product
            line = start + 1
            worksheet.merge_range(("C%d:H%d" % (line, line)), product['product'])

            line += 1
            option_str = get_product_type_name(product['product']) + ' '
            for option in product['extra']['product_options_value']:
                option_str += option['option_name'] + ': ' + option['variant_name'] + ' '
            option_str += '数量: ' + product['amount']
            worksheet.merge_range(("C%d:H%d" % (line, line)), option_str)

            line += 1
            worksheet.merge_range(("C%d:H%d" % (line, line)), '')

            line += 1
            worksheet.merge_range(("F%d:H%d" % (line, line)), order.s_firstname + ' ' + order.s_lastname)

            line += 1
            worksheet.merge_range(("F%d:H%d" % (line, line)), order.s_address)

            line += 1
            worksheet.merge_range(("F%d:H%d" % (line, line)), order.s_address_2)

            line += 1
            worksheet.merge_range(("F%d:H%d" % (line, line)), order.s_city + ' ' + order.s_state + ' ' + order.s_zipcode)

            line += 1
            worksheet.merge_range(("F%d:H%d" % (line, line)), order.s_country)

            line += 1
            worksheet.merge_range(("F%d:H%d" % (line, line)), order.s_phone)

            line += 1
            worksheet.merge_range(("F%d:H%d" % (line, line)), '')

            # image
            url = order.product_groups[0]['products'][item_id]['main_pair']['icon']['image_path']
            image_origin_x = int(order.product_groups[0]['products'][item_id]['main_pair']['icon']['image_x'])
            image_origin_y = int(order.product_groups[0]['products'][item_id]['main_pair']['icon']['image_y'])
            import urllib, io
            image_data = io.BytesIO(urllib.request.urlopen(urllib.request.Request(url)).read())
            worksheet.merge_range(("C%d:E%d" % (start+4, end+1)), '')
            worksheet.insert_image(("C%d" % (start+4)), url, {'image_data': image_data, 'x_scale': 180/image_origin_x, 'y_scale': 180/image_origin_y})

            # supply
            merge_lines_data(worksheet, start, 8, interval, '', merge_format)
            # supply price
            merge_lines_data(worksheet, start, 9, interval, '', merge_format)
            # Express
            merge_lines_data(worksheet, start, 10, interval, order.shipping_method, merge_format)
            # Tracking No.
            merge_lines_data(worksheet, start, 11, interval, order.tracking_no, merge_format)
            # Express price
            merge_lines_data(worksheet, start, 12, interval, '', merge_format)
            # Ship Time
            merge_lines_data(worksheet, start, 13, interval, order.ship_time, merge_format)
            # product memo
            merge_lines_data(worksheet, start, 14, interval, '', merge_format)
            # order memo
            merge_lines_data(worksheet, start, 15, interval, '', merge_format)

            start += interval
            end = start + interval - 1

    workbook.close()

    # update download time
    db.session.commit()

    # https://stackoverflow.com/questions/50086585/download-file-from-root-directory-using-flask
    return send_file(xlsx_filename, as_attachment=True)


# /order/index
@bp.route('/index', methods=('GET', 'POST'))
def order_index():
    channels = Channel.query.all()  # get all channels
    events = Event.query.all()  # get all events
    orders_query = db.session.query(Order)  # create Order Query
    active_channel = None
    active_status = None
    active_event = None
    keyword = None

    if request.method == 'POST':  # search or batch
        # get arguments
        keyword = request.form.get('keyword')
        channel_id = request.form.get('channel_id')
        status = request.form.get('status')
        event_id = request.form.get('event')
        status_to = request.form.get('status_to')

        # set Order query with filter
        # https://stackoverflow.com/questions/37336520/sqlalchemy-dynamic-filter
        # search
        if channel_id:  # Add conditions
            orders_query = orders_query.filter(Order.channel_id == channel_id)
            active_channel = Channel.query.filter(
                Channel.id == channel_id).first()
        if status:  # status filter
            orders_query = orders_query.filter(Order.status == status)
            active_status = status
        if event_id:  # event filter
            # https://stackoverflow.com/questions/6474989/sqlalchemy-filter-by-membership-in-at-least-one-many-to-many-related-table
            orders_query = orders_query.filter(
                Order.events.any(Event.id.in_([event_id])))
            active_event = Event.query.filter(Event.id == event_id).first()
        if keyword:  # search keyword
            orders_query = orders_query.filter(
                or_(Order.order_id == keyword, Order.email == keyword))
        # batch
        if status_to:  # change order status
            order_ids = request.form.get('order_ids')
            if order_ids:
                order_ids = [int(x) for x in order_ids.split(',')
                             ]  # change string to integer
            # https://stackoverflow.com/questions/15267755/query-for-multiple-values-at-once
            batching_orders = Order.query.filter(Order.id.in_(order_ids))
            for order in batching_orders:
                order.status = status_to
            db.session.commit()

    orders = orders_query.all()  # execute Query

    return render_template('order/index.html',
                           active_page="order",
                           orders=orders,
                           events=events,
                           channels=channels,
                           active_channel=active_channel,
                           active_status=active_status,
                           active_event=active_event,
                           keyword=keyword)


# /order/detail
@bp.route('/detail/<string:order_id>', methods=('GET', ))
def order_detail(order_id: str):
    order = Order.query.filter(Order.id == order_id).first()

    return render_template('/order/detail.html', order=order)


# /order/sync
@bp.route('/sync', methods=('GET', ))
def order_sync():
    order_keys = get_model_keys(Order)

    for channel in Channel.query.all():
        api = Cscart(channel.website_url, channel.email, channel.api_key)
        response = api.get_orders()
        for o in response['orders']:
            order_id = o['order_id']
            order_remote = api.get_order(order_id)
            order_local = Order.query.filter(
                Order.order_id == order_id).first()

            order_remote['shipping_method'] = order_remote['shipping'][0]['shipping']  # shipping method

            # choose keys in order_remote
            order_remote = {k: v for k, v in order_remote.items() if k in order_keys}
            order_remote['channel'] = channel  # add foreign key

            # timestamp to datetime
            # https://stackoverflow.com/questions/3682748/converting-unix-timestamp-string-to-readable-date
            order_remote['timestamp'] = datetime.utcfromtimestamp(int(order_remote['timestamp']))

            if order_local:
                for k, v in order_remote.items():
                    setattr(order_local, k, v)
            else:
                # https://stackoverflow.com/questions/31750441/generalised-insert-into-sqlalchemy-using-dictionary/31756880
                order = Order(**order_remote)
                db.session.add(order)
            current_app.logger.debug("Channel %s:Order %s synchronized!" % (channel.name, order_id))
        db.session.commit()
        current_app.logger.info("Channel #%s synchronized!" % (channel.name))

    flash('Synchronization Completed!', 'success')
    return redirect(url_for('order.order_index'))


# /order/updat_estatus/<str:channel_id>/<str:order_id>/<str:status>
@bp.route(
    '/update_status/<string:channel_id>/<string:order_id>/<string:status>',
    methods=('GET', ))
def order_update_status(channel_id: str, order_id: str, status: str):
    order = Order.query.filter(Order.id == order_id).filter(
        Order.channel_id == channel_id).first()
    channel = order.channel

    #api = Cscart(channel.website_url, channel.email, channel.api_key)
    #response = api.update_order_status(str(order.order_id), status)

    order.status = status
    db.session.commit()

    flash('Update Order Status Completed', 'success')
    current_app.logger.debug("Channel #%s Order #%s status change to %s" %
                             (channel.name, order.id, order.status))

    return redirect(url_for('order.order_index'))
